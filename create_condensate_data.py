import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# สร้างข้อมูลตัวอย่าง (1 มกราคม 2026 ถึง 31 ธันวาคม 2026)
start_date = datetime(2026, 1, 1)
dates = [(start_date + timedelta(days=x)).strftime('%Y-%m-%d') for x in range(366)]
data = {
    'วันที่': dates,
    'เวลา': [f"{random.randint(0,23):02d}:{random.randint(0,59):02d}" for _ in dates],
    'ปริมาณน้ำควบแน่น (ลิตร)': [round(random.uniform(50, 500), 2) for _ in dates],
    'อุณหภูมิ (°C)': [round(random.uniform(60, 100), 1) for _ in dates],
    'ความดัน (bar)': [round(random.uniform(3, 8), 2) for _ in dates],
    'คุณภาพน้ำ (TDS)': [round(random.uniform(500, 2000), 0) for _ in dates],
    'หมายเหตุ': ['ปกติ' if random.random() > 0.2 else 'ผิดปกติ' for _ in dates]
}

df = pd.DataFrame(data)

# สร้าง workbook ใหม่
wb = Workbook()
ws = wb.active
ws.title = '%condensate'

# กำหนดคำสั่งหัวตาราง
headers = df.columns
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# เพิ่มข้อมูล
for row_num, row_data in enumerate(df.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # เพิ่มเส้นขอบ
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = border

# ปรับความกว้างของคอลัมน์
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

# บันทึกไฟล์
wb.save('condensate_data.xlsx')
print("✅ สร้างไฟล์ condensate_data.xlsx สำเร็จ")
print(f"📊 ข้อมูล: {len(df)} บันทึก ({len(df)} วัน)")
print(f"📅 ช่วงวันที่: {df['วันที่'].iloc[0]} ถึง {df['วันที่'].iloc[-1]}")
